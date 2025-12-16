import shutil
import os
import openpyxl
from openpyxl.utils import get_column_letter
from typing import Optional, Callable
from app.models import MinerType, WarrantyDetails
from app.api_clients import check_warranty
from app.serial_parser import is_serial_number

HEADER_WARRANTY = "Warranty Expires"
HEADER_MINER_TYPE = "Miner Type"

def find_serial_column(sheet, max_rows=10) -> int:
    """
    Scans rows to find the serial column.
    
    Logic ported from Java (lines 118-200):
    - Iterates through the sheet row by row.
    - If serialColumn is not yet found:
        - Checks every cell in the current row.
        - If cell content matches serial pattern -> Found! Set serialColumn.
        - OR if cell content contains "serial" -> Found! Set serialColumn.
    - Once found, this function would theoretically return the column index.
    
    However, the Java code does this INSIDE the main loop. 
    To adapt for our needs, we will scan the first 'max_rows' rows (or more) 
    to find the column index before processing the file.
    """
    print(f"DEBUG: Scanning first {max_rows} rows for serial column...")
    
    for row in range(1, min(max_rows + 50, sheet.max_row + 1)): # Increased scan depth
        for col in range(1, sheet.max_column + 1):
            cell_val = sheet.cell(row=row, column=col).value
            if cell_val and isinstance(cell_val, str):
                s_cell = cell_val.strip()
                
                # Check 1: Regex Pattern Match
                if is_serial_number(s_cell):
                    print(f"DEBUG: Found serial using PATTERN in row {row}, col {col}: '{s_cell}'")
                    return col
                
                # Check 2: Header Keyword Match
                if "serial" in s_cell.lower():
                    print(f"DEBUG: Found serial using KEYWORD in row {row}, col {col}: '{s_cell}'")
                    return col
                    
    print("DEBUG: Serial column detection failed.")
    return -1

def process_file_logic(input_path: str, output_path: str, miner_type: MinerType, progress_callback: Callable[[int, int, str], None] = None):
    """
    Main logic to process the Excel file.
    Resumes if output_path exists and seems valid.
    """
    # 1. Prepare Output File
    resume = False
    if os.path.exists(output_path):
        # Check if we can resume
        try:
            wb = openpyxl.load_workbook(output_path)
            sheet = wb.active
            # Check for headers
            last_col = sheet.max_column
            header_w = sheet.cell(row=1, column=last_col-1).value
            header_m = sheet.cell(row=1, column=last_col).value
            
            # Simple check if last columns are ours. 
            # Note: in auto mode we add 2 cols. In others just 1?
            # Java adds "Miner Type" ONLY if MinerType.AUTO.
            # "Warranty Expires" is always added.
            
            # Let's verify headers matching logic
            # Actually, simpler resume logic: 
            # If explicit output file exists, assume it is the working file.
            resume = True
        except Exception:
            # If failed to load, overwrite
            resume = False
    
    if not resume:
        shutil.copy2(input_path, output_path)
        wb = openpyxl.load_workbook(output_path)
        sheet = wb.active
    
    # 2. Setup Columns
    # We need to find or create the detail columns
    warranty_col_idx = -1
    type_col_idx = -1
    
    # Check headers
    headers_row = 1
    max_col = sheet.max_column
    
    # Try to find existing headers
    for c in range(1, max_col + 1):
        val = sheet.cell(row=headers_row, column=c).value
        if val == HEADER_WARRANTY:
            warranty_col_idx = c
        if val == HEADER_MINER_TYPE:
            type_col_idx = c
            
    if warranty_col_idx == -1:
        # Create them
        warranty_col_idx = max_col + 1
        sheet.cell(row=headers_row, column=warranty_col_idx).value = HEADER_WARRANTY
        if miner_type == MinerType.AUTO:
            type_col_idx = warranty_col_idx + 1
            sheet.cell(row=headers_row, column=type_col_idx).value = HEADER_MINER_TYPE
        
        # Save initializing
        wb.save(output_path)

    # 3. Find Serial Column
    serial_col_idx = find_serial_column(sheet)
    if serial_col_idx == -1:
        print("Could not detect serial number column.")
        return

    # 4. Iterate Rows
    total_rows = sheet.max_row
    
    # We create a new client session or just loop
    for r in range(2, total_rows + 1):
        # Check if already done
        w_cell = sheet.cell(row=r, column=warranty_col_idx)
        if w_cell.value:
            # Already has data, skip (Resume logic)
            continue
        
        serial_cell = sheet.cell(row=r, column=serial_col_idx)
        serial = serial_cell.value
        if not serial or not isinstance(serial, str):
            continue
            
        serial = serial.strip()
        if not serial:
            continue
        
        # Update Progress
        if progress_callback:
            progress_callback(r, total_rows, f"Checking {serial}...")
            
        # Lookup
        details = check_warranty(serial, miner_type)
        
        # Write
        if details:
            w_cell.value = details.warranty_date
            if miner_type == MinerType.AUTO and type_col_idx != -1:
                t_cell = sheet.cell(row=r, column=type_col_idx)
                t_cell.value = details.type.value
        
        # Save every 5 rows to be safe and support resume
        if r % 5 == 0:
            wb.save(output_path)

    # Final Save
    wb.save(output_path)
    if progress_callback:
        progress_callback(total_rows, total_rows, "Done!")

